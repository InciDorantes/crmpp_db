from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from PPRegister.models import UserProfile, programas_p, permisos_pps
from django.core.mail import send_mail
from django.core.serializers.json import DjangoJSONEncoder
from PPRegister.models import PoblacionObjetivo, BienServicio
from PPRegister.models import APED
from PPRegister.models import FinMir, PropositoMir, IndicadorPropositoMir, ComponenteMir, IndicadorComponenteMir,ActividadMir, IndicadorActividadMir
from PPRegister.models import FichaIndicador
from PPRegister.models import FormatoUno, FormatoDos, FormatoTres, FormatoCuatro, Nodo, Conexion, NodoAO, ConexionAO
from PPRegister.models import Municipios
from PPRegister.models import FuenteFinanciamiento, Formato9, ComponenteMir, FormatoQuince
from django.db.models import Case, When, Value, IntegerField
#LOGEO ENTRADA A HOME
from django.shortcuts import redirect
from django.contrib import messages
from django.urls import reverse
from django.forms.models import model_to_dict
from django.utils.safestring import mark_safe
from django.http import HttpResponse
from django.http import JsonResponse, HttpResponseRedirect

from .utils.excelfunction import (
    Llenar_f5,
    Llenar_f7,
    Llenar_f8,
    crear_arbol,
    crear_encabezado,
    formato_celda_cuerpo,
    formato_encabezado_tablas,
    formato_esqueleto_fichas,
    llenar_f1,
    llenar_f3,
    llenar_f2,
    llenar_f4,
    llenar_f6,
    Llenar_f7,
    llenar_f9,
    llenar_f10,
    llenar_f13,
    llenar_f14,
    llenar_ficha_en_hoja,
    merge_columns_flexible,
    merge_dos_filas,
    merge_filas,
    llenar_f17,
    llenar_f15,
    llenar_f16,
    llenar_f12
)

from .utils.structurefuncions import (
    estructura_AO,
    estructura_AP,
    estructuraF9,
    estructura_formato6,
    estructura_formatoDoce,
    estructura_formatoTrece,
    estructura_formato17,
    estructura_poblacion_obj,
    estructura_usuarios,
    estructurar_fichas,
    generar_arbol_objetivos,
    generar_data_mir,
    generar_estructura_actividades,
    generar_estructura_aped,
    generar_estructura_componentes,
    obtener_ficha_fin_estructurada,
    obtener_ficha_proposito_estructurada,
    estructura_formatoCatorce_envio,
    estructura_formato15,
    generar_estructura_f16
)

from .utils.savefunctions import (
    guardarArbolObjetivosNuevo,
    guardarArbolProblemaNuevo,
    guardarFormatoUno,
    guardarFormatoDos,
    guardarFormatoTres,
    guardarFormatoCuatro,
    guardar_poblacion_objetivo,
    guardarFormatoSeis,
    guardarAPED,
    guardar_bienes_servicios,
    guardarFormatoNueve,
    guardar_datos_mir,
    guardar_ficha_fin,
    guardar_fichas_componentes,
    guardar_fichas_componentes_actividad,
    guardar_fichas_proposito,
    guardarFormatoDoce,
    guardarFormatoTrece,
    guardarFormatoCatorce,
    guardarFormatoQuince,
    guardarFormatoDieciseis,
    guardarFormatoDieciciete
)

import requests
import os
import json
import openpyxl
import io
from openpyxl.drawing.image import Image as XLImage
from PIL import Image  
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.merge import MergedCell
from openpyxl import Workbook
from io import BytesIO
import base64
from reportlab.pdfgen import canvas
#from xhtml2pdf import pisa
import base64
import tempfile
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from django.http import FileResponse
from django.contrib.auth.models import User

from io import BytesIO
from openpyxl import load_workbook
from django.utils.html import mark_safe
from django.shortcuts import render
from PIL import Image
import base64
from .models import programas_p


def home(request):
    if not request.user.is_authenticated:
        return render(request, 'home.html', {'is_authenticated': False})

    # 1) Filtrado inicial de programas
    if request.user.is_superuser:  # o cualquier otra condición para ver todos
        lista_pps = programas_p.objects.all()
    else:
        lista_pps = programas_p.objects.filter(
            permisos_pps__users=request.user
        ).distinct()
    
    user = request.user
    tipo_usuario = user.userprofile.role

    # 2) Manejo del formulario de guardar programa
    if 'guardar_programa_home' in request.POST:
        nombre_pp = request.POST.get('nombre_pp')
        num_pp = request.POST.get('num_pp')
        num_pp = int(num_pp) if num_pp and num_pp.isdigit() else 0
        usuarios_ids = request.POST.getlist('usuarios')

        # Validaciones de unicidad
        if programas_p.objects.filter(nombre_pp=nombre_pp).exists():
            messages.error(request, "El nombre de pp ya existe")
            return render(request, 'home.html', {
                'is_authenticated': True,
                'nombre_pp': nombre_pp,
                'num_pp': num_pp,
                'usuarios': User.objects.all(),
                'lista_pps': lista_pps,
            })
        if num_pp != 0 and programas_p.objects.filter(num_pp=num_pp).exists():
            messages.error(request, "El número de pp ya existe")
            return render(request, 'home.html', {
                'is_authenticated': True,
                'nombre_pp': nombre_pp,
                'num_pp': num_pp,
                'usuarios': User.objects.all(),
                'lista_pps': lista_pps,
            })

        # Crear el programa
        nuevo_pp = programas_p.objects.create(
            nombre_pp=nombre_pp,
            num_pp=num_pp,
            user_id = user.id
        )

        # Permisos: el creador como editor
        permisos_pps.objects.create(
            id_pp=nuevo_pp,
            users=request.user,
            editor=True
        )
        # Permisos: los demás usuarios como no editores
        for user_id in usuarios_ids:
            usuario = User.objects.get(id=user_id)
            if usuario != request.user:
                permisos_pps.objects.create(
                    id_pp=nuevo_pp,
                    users=usuario,
                    editor=False
                )

        # Guardar en sesión para mostrar después del redirect
        request.session['nombre_pp'] = nombre_pp
        request.session['num_pp'] = num_pp
        request.session['usuarios_seleccionados'] = usuarios_ids
        messages.success(request, "Programa y usuarios guardados correctamente!")
        return redirect('home')
    
    if 'guardar_numpp' in request.POST:
        id_pp = request.POST.get('id_pp')
        num_pp = request.POST.get('num_pp_superuser')

        if id_pp and num_pp and num_pp.isdigit():
            try:
                programa = programas_p.objects.get(id_pp=id_pp)
                if programas_p.objects.filter(num_pp=num_pp).exclude(id_pp=id_pp).exists():
                    messages.error(request, "Este número de pp ya está en uso.")
                else:
                    programa.num_pp = int(num_pp)
                    programa.save()
                    messages.success(request, "Número de pp guardado correctamente.")
            except programas_p.DoesNotExist:
                messages.error(request, "Programa no encontrado.")
        else:
            messages.error(request, "Datos inválidos.")

        return redirect('home')
    # 3) Recuperar datos de sesión tras redirect
    nombre_pp = request.session.pop('nombre_pp', None)
    num_pp = request.session.pop('num_pp', None)
    usuarios_seleccionados = request.session.pop('usuarios_seleccionados', [])
    
    # 4) Construir programas_data con el flag es_editor
    programas_data = []
    for pp in lista_pps:
        es_editor = permisos_pps.objects.filter(
            id_pp=pp,
            users=request.user,
            editor=True
        ).exists()
        programas_data.append({
            'pp': pp,
            'es_editor': es_editor
        })

    #5) datos de los usuarios
    perfil = request.user.userprofile
    clasificacion = perfil.clasificacion

    if tipo_usuario == "superuser":
        user_profiles = UserProfile.objects.all()
    else:
        user_profiles = UserProfile.objects.filter(clasificacion=clasificacion)

    # 6) Renderizar pasando programas_data
    return render(request, 'home.html', {
        'is_authenticated': True,
        'nombre_pp': nombre_pp,
        'num_pp': num_pp,
        'usuarios_seleccionados': usuarios_seleccionados,
        'usuarios': User.objects.all(),
        'user_profiles': user_profiles,
        'programas_data': programas_data,
        'tipo_usuario': tipo_usuario,
    })

def registro(request, id_pp):
    try:
        # Obtener el programa basado en el id_pp
        pp = programas_p.objects.get(id_pp=id_pp)
        id = pp.id_pp
        user = pp.user
        tipo_usuario = user.userprofile.role
        #Variables
        arbol_ap ={}
        arbol_obj ={}
        bien_existente= BienServicio.objects.filter(id_pp=pp).exists()
        fin_existente = FinMir.objects.filter(id_pp=pp).exists()
        data ={}
        estructura_aped = {}
        bienes_servicios_guardados ={}
        data_mir ={}
        poblaciones_dict = {}  
        datos_ficha_fin= {}
        datos_ficha_proposito= {}
        datos_ficha_componente = {}
        datos_ficha_actividad = {}
        formato4 ={}

        #Formato 14
        formato14_tabla = estructura_formatoCatorce_envio(id)
        #Formato9
        permisos = permisos_pps.objects.filter(id_pp=id_pp).select_related('users')
        usuarios_con_datos = estructuraF9(permisos, pp.id_pp)

        #usuarios
        usuarios_json =estructura_usuarios(user, tipo_usuario)
        if Nodo.objects.filter(id_pp__id_pp=id).exists():
            arbol_ap = estructura_AP(id)

        #si arbol de objetivos no existe enviar sugerencia
        if  NodoAO.objects.filter(id_pp__id_pp=id).count() == 0:
            arbol_obj = generar_arbol_objetivos (estructura_AP(id_pp))
            #si arbol de objetivos sí existe enviar guardado
        elif Nodo.objects.filter(id_pp__id_pp=id).exists() and NodoAO.objects.filter(id_pp__id_pp=id).exists():
            arbol_obj = estructura_AO(id)

        if bien_existente:
            bienes_guardados = BienServicio.objects.filter(id_pp=pp)
            bienes_lista = [model_to_dict(bien) for bien in bienes_guardados]
            bienes_servicios_guardados = mark_safe(json.dumps(bienes_lista))
        
        if fin_existente:
            datamir = generar_data_mir(pp)
            data_mir = json.dumps(datamir, cls=DjangoJSONEncoder)

        if PoblacionObjetivo.objects.filter(id_pp=id).exists():
            poblaciones_dict = estructura_poblacion_obj(id)

        if FichaIndicador.objects.filter(id_pp_id=id,tipo_ficha='fin').exists():
            datos_ficha_fin = obtener_ficha_fin_estructurada(id)

        if FichaIndicador.objects.filter(id_pp_id=id,tipo_ficha='proposito').exists():
            datos_ficha_proposito = obtener_ficha_proposito_estructurada(id)
        
        if APED.objects.filter(id_pp_id=id_pp).exists():
            estructura_aped = generar_estructura_aped(pp.id_pp)
        # Procesar componentes
        if FichaIndicador.objects.filter(id_pp_id=id, tipo_ficha='componente').exists():
            resultado_componentes = generar_estructura_componentes(id)
            if resultado_componentes['status'] == 'ok':
                datos_ficha_componente = resultado_componentes['data']
            else:
                print(f"Error al generar componentes: {resultado_componentes['message']}")
        
        # Procesar actividades
        if FichaIndicador.objects.filter(id_pp_id=id, tipo_ficha='actividad').exists():
            resultado_actividades = generar_estructura_actividades(id)
            if resultado_actividades['status'] == 'ok':
                datos_ficha_actividad = resultado_actividades['data']
            else:
                print(f"Error al generar actividades: {resultado_actividades['message']}")
        if FormatoCuatro.objects.filter(id_pp=id):
            formato4 = FormatoCuatro.objects.get(id_pp=pp)

        componentes_f16 = generar_estructura_f16(id)
        datos_f15 = estructura_formato15(id)

        #recepción de datos 
        if request.method == 'POST':
            # Determinar si es AJAX
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            try:
                if is_ajax:
                    # Procesamiento para peticiones AJAX
                    data = json.loads(request.body)
                    action = data.get('action')
                    if action =='guardar_arbol_problemas_nuevo':
                         datos = data.get('datos_arbol', [])
                         return guardarArbolProblemaNuevo(datos, pp.id_pp)
                    
                    elif action =='guardar_arbol_objetivos_nuevo':
                         datos = data.get('datos_arbol', [])
                         return guardarArbolObjetivosNuevo(datos, pp.id_pp)
                    
                    elif action =='guardar_formato_uno':
                         return guardarFormatoUno(data, pp.id_pp)
                    
                    elif action =='guardar_formato_dos':
                         return guardarFormatoDos(data, pp.id_pp)
                    
                    elif action =='guardar_formato_tres':
                         return guardarFormatoTres(data, pp.id_pp)
                    
                    elif action =='guardar_formato_cuatro':
                         return guardarFormatoCuatro(data, pp.id_pp)
                    
                    elif action =='guardar_poblacion_objetivo':
                        return guardar_poblacion_objetivo(pp, data)
                    
                    elif action =='guardar_formato_seis':
                        datos =data.get('data', [])
                        return guardarFormatoSeis(datos,id_pp)
                    
                    elif action == 'guardar_aped':
                        datos = data.get('data',[]) 
                        return guardarAPED(datos, pp.id_pp)
                    
                    elif action == 'guardar_bienes_servicios':
                        return guardar_bienes_servicios(request, pp, data)
                    
                    elif action == 'guardar_formato_nueve':
                        datos = data.get('data',[])                        
                        return guardarFormatoNueve(datos, pp.id_pp)
                    
                    
                    elif action == 'guardar_mir':
                        return guardar_datos_mir(request, pp, data, is_ajax=True)
                    
                    elif action == 'guardar_ficha_fin':
                        return guardar_ficha_fin(pp, data)
                    elif action == 'guardar_ficha_proposito':
                        return guardar_fichas_proposito(pp, data, request)
                    elif action == 'guardar_ficha_componente':
                        componentes = data.get('data', [])
                        return guardar_fichas_componentes(pp.id_pp, componentes)
                    elif action == 'guardar_ficha_actividad':
                         return guardar_fichas_componentes_actividad(pp.id_pp, data.get('data', []))
                    
                    elif action =='guardar_formato_doce':
                        datos =data.get('data', [])
                        return guardarFormatoDoce(datos,id_pp)
                    
                    elif action =='guardar_formato_trece':
                        datos =data.get('data', [])
                        return guardarFormatoTrece(datos,pp.id_pp)
                    
                    elif action =='guardar_formato_catorce':
                        datos =data.get('data', [])
                        return guardarFormatoCatorce(datos, pp.id_pp)
                    
                    elif action =='guardar_formato_quince':
                        datos =data.get('data', [])
                        return  guardarFormatoQuince(datos, pp.id_pp)
                    
                    elif action =='guardar_formato_dieciseis':
                        datos =data.get('data', [])
                        return guardarFormatoDieciseis(datos, pp.id_pp)
                    
                    elif action =='guardar_formato_dieciciete':
                        datos =data.get('data', [])
                        return guardarFormatoDieciciete(datos,pp.id_pp)

                else:
                    if 'reset-ap' in request.POST:
                        print("se reseteo")

                        # Redirigir a la vista para mostrar el estado actualizado
                        return render(request, 'registro.html', {
                            'is_authenticated': True,
                            'pp': pp,
                            'usuarios': User.objects.all(),
                            'tree_data_json': arbol_ap,
                            'tree_data_ao_json': arbol_obj,
                            'problema_central_existente': Nodo.objects.filter(id_pp__id_pp=id).exists(),
                            'solucion_problema_existente': NodoAO.objects.filter(id_pp__id_pp=id).exists(),
                            'PoblacionObjetivo':PoblacionObjetivo.objects.filter(id_pp=pp),
                            'poblaciones': poblaciones_dict,
                            'estructura_aped': estructura_aped,
                            'aped_existe':APED.objects.filter(id_pp_id=id_pp).exists(),
                            'bienes_servicios_guardados': bienes_servicios_guardados,
                            'datos_mir_json':data_mir,
                            'datos_ficha_fin': json.dumps(datos_ficha_fin, cls=DjangoJSONEncoder),
                            'datos_fichas_proposito': json.dumps(datos_ficha_proposito,cls=DjangoJSONEncoder),
                            'datos_fichas_componente': json.dumps(datos_ficha_componente, cls=DjangoJSONEncoder),
                            'datos_fichas_actividad': json.dumps(datos_ficha_actividad, cls=DjangoJSONEncoder),
                            'formatouno': FormatoUno.objects.filter(id_pp=pp),
                            'usuarios_con_datos': usuarios_con_datos,
                            "usuarios_con_datos_json": usuarios_json,
                            'municipios': Municipios.objects.all(),
                            'ff': FuenteFinanciamiento.objects.all(),
                            'formato6':estructura_formato6(pp.id_pp), 
                            'formato12': estructura_formatoDoce(pp.id_pp),
                            'componentes': componentes_f16,
                            'data_formato14': formato14_tabla,
                            'datos_f13': estructura_formatoTrece(pp.id_pp),
                            'datos_f15': datos_f15,
                            'datos_f17':estructura_formato17(pp.id_pp)
                        })
        
                    else:
                        messages.error(request, 'Acción no reconocida')
                        return HttpResponseRedirect(reverse('registro', args=[id_pp]))
            
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
            
            except Exception as e:
                if is_ajax:
                    return JsonResponse({'error': str(e)}, status=500)
                else:
                    messages.error(request, f'Error: {str(e)}')
                    return HttpResponseRedirect(reverse('registro', args=[id_pp]))
          
        else:
            return render(request, 'registro.html', {
                'is_authenticated': True,
                'usuarios': User.objects.all(),
                'pp':pp,
                'tree_data_json': arbol_ap,
                'arbol_objetivos': arbol_obj,
                'problema_central_existente': Nodo.objects.filter(id_pp__id_pp=id).exists(),
                'solucion_problema_existente': NodoAO.objects.filter(id_pp__id_pp=id).exists(),
                'PoblacionObjetivo':PoblacionObjetivo.objects.filter(id_pp=pp),
                'poblaciones': poblaciones_dict,
                'aped_existe':APED.objects.filter(id_pp_id=id_pp).exists(),
                'estructura_aped': estructura_aped,
                'bienes_servicios_guardados': bienes_servicios_guardados,
                'datos_mir_json':data_mir,
                'datos_ficha_fin': json.dumps(datos_ficha_fin, cls=DjangoJSONEncoder),
                'datos_fichas_proposito': json.dumps(datos_ficha_proposito,cls=DjangoJSONEncoder),
                'datos_fichas_componente': json.dumps(datos_ficha_componente, cls=DjangoJSONEncoder),
                'datos_fichas_actividad': json.dumps(datos_ficha_actividad, cls=DjangoJSONEncoder),
                'formato1_list': FormatoUno.objects.filter(id_pp=pp),
                'formato2_list': FormatoDos.objects.filter(id_pp=pp),
                'formato3_list': FormatoTres.objects.filter(id_pp=pp),
                'formato4': formato4,
                'usuarios_con_datos': usuarios_con_datos,
                "usuarios_json": usuarios_json,
                'municipios': Municipios.objects.all(),
                'ff': FuenteFinanciamiento.objects.all(),
                'formato6':estructura_formato6(pp.id_pp),
                'formato12': estructura_formatoDoce(pp.id_pp),
                'componentes': componentes_f16,
                'data_formato14': formato14_tabla,
                'datos_f13': estructura_formatoTrece(pp.id_pp),
                'datos_f15': datos_f15,
                'datos_f17':estructura_formato17(pp.id_pp)
            })    

    except programas_p.DoesNotExist:
        # Si no se encuentra el programa, redirigir o mostrar un error
        return redirect('home')  # O puedes mostrar una página de error

def visualizar(request, id_pp):
    if not request.user.is_authenticated:
        return render(request, 'home.html', {'is_authenticated': False})

    programa = programas_p.objects.get(id_pp=id_pp)
    wb = view_excel(programa.id_pp)  # tu función que genera el workbook

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb = load_workbook(filename=output, data_only=True)

    hojas_html = []
    tab_buttons = []

    for i, sheet in enumerate(wb.worksheets):
        div_id = f"sheet_{i}"
        tab_buttons.append(
            f"<button type='button' class='tab-button' data-target='{div_id}'>{sheet.title}</button>"
        )

        html = f"<div id='{div_id}' class='sheet-content' style='display: {'block' if i == 0 else 'none'};'>"

        if sheet.title.lower().startswith("árbol"):
            if hasattr(sheet, "_images") and len(sheet._images) >= 2:
                # Encontrar la imagen anclada en la fila más baja (si hay 2 o más)
                imagenes_con_fila = []
                for img in sheet._images:
                    try:
                        fila = img.anchor._from.row
                    except AttributeError:
                        fila = -1
                    imagenes_con_fila.append((fila, img))

                imagenes_con_fila.sort(key=lambda x: x[0], reverse=True)
                fila_mas_baja, img = imagenes_con_fila[0]

                image_data = img._data()
                image_bytes = BytesIO(image_data)
                pil_img = Image.open(image_bytes)

                buffer = BytesIO()
                pil_img.save(buffer, format="PNG")
                img_str = base64.b64encode(buffer.getvalue()).decode("utf-8")

                html += f"<img src='data:image/png;base64,{img_str}' alt='Imagen embebida' style='max-width:100%; height:auto; margin-bottom:10px;'/>"
            else:
                # 0 o 1 imagen => mostrar en blanco (no mostrar imagen)
                html += "<p><em>No hay imágenes para mostrar (menos de 2 imágenes en esta hoja).</em></p>"
        else:
            # Construir tabla con estilos para primera y segunda fila
            html += "<table class='excel-table' border='1' style='border-collapse:collapse; width:100%;'>"
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx == 0:
                    style = "background-color:#5d1f3f; color:white;"
                elif idx == 1:
                    style = "background-color:#c2995cd7;"
                else:
                    style = ""

                html += f"<tr style='{style}'>" + "".join(
                    f"<td>{cell if cell is not None else ''}</td>" for cell in row
                ) + "</tr>"
            html += "</table>"

        html += "</div>"
        hojas_html.append(html)

    return render(request, 'visualizar.html', {
        'excel_html': mark_safe("".join(tab_buttons) + "".join(hojas_html)),
        'is_authenticated': True
    })


#LOGIN Y LOGOUT
def login_view(request):
   if request.method == "POST":
        username = request.POST.get("username")  # Captura el usuario
        password = request.POST.get("password")  # Captura la contraseña

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("home")  # Redirige al home si el login es exitoso
        else:
            messages.error(request, "Usuario o contraseña incorrectos") 
   return render(request, 'login.html', {})

def logout_view(request):
    logout(request)
    return redirect("login")

def view_excel(id_pp):
    programa = programas_p.objects.get(id_pp=id_pp)
    nombre = programa.nombre_pp
    id = programa.id_pp
    entidad = programa.user.userprofile.siglas

    #arboles
    arbol_problemas= {'nodos':[], 'conexiones':[]}
    if Nodo.objects.filter(id_pp__id_pp=id).exists():
            arbol_problemas = estructura_AP(id)
    arbol_objetivos = {'nodos':[], 'conexiones':[]}
    if Nodo.objects.filter(id_pp__id_pp=id).exists() and NodoAO.objects.filter(id_pp__id_pp=id).exists():
            arbol_objetivos = estructura_AO(id)

    test_formato1 = {}

    test_formato2 = {}

    test_formato3 = {}

    test_formato4 = {}

    test_formato5 ={}
    if PoblacionObjetivo.objects.filter(id_pp=id).exists():
        test_formato5 = estructura_poblacion_obj(id)
    
    test_formato6 = {}
        

    test_formato8 ={}
    if BienServicio.objects.filter(id_pp=id).exists():
        bienes_guardados = BienServicio.objects.filter(id_pp=id)
        test_formato8 = [model_to_dict(bien) for bien in bienes_guardados]    

    test_formato9 = {}
    
    test_formato10= {}
    if FichaIndicador.objects.filter(id_pp_id=id,tipo_ficha='fin').exists():
        test_formato10 = generar_data_mir(id)
    
    test_formato11 = {}
    test_formato12 = {}
    test_formato13 = {}
    test_formato14 = {}
    test_formato15 = {}
    test_formato16 = {}
    test_formato17 = {}

    wb = Workbook()
    titulos_y_columnas = { 
                'Formato 1': ['. Documentación de buenas prácticas y programas similares', 8],
                'Formato 2': ['. Vinculación con otros programas implementados en el estado', 11],
                'Formato 3': ['. Identificación de involucrados', 5],
                'Formato 4': ['. Criterios para la focalización de la población objetivo', 3],
                'Árbol de Problemas': 14,
                'Árbol de Objetivos': 14,
                'Formato 5': ['. Identificación y cuantificación de la población objetivo', 9],
                'Formato 6': ['. Cobertura geográfica', 12],
                'Formato 7': ['. Alineación con la planeación del desarrollo', 2],
                'Formato 8': ['. Características de los bienes y/o servicios del programa', 4],
                'Formato 9': ['. Corresponsabilidad  interinstitucional', 8],
                'Formato 10': ['. Matriz de Indicadores para Resultados (MIR)', 5],
                'Formato 11': ['. Ficha de indicadores', 4],
                'Formato 12': ['. Fuentes de Información', 8], 
                'Formato 13': ['. Informes de desempeño', 4],
                'Formato 14': ['. Marco de resultados en el mediano plazo', 10], 
                'Formato 15': ['. Programación de la atención a la población objetivo en el mediano plazo', 8],
                'Formato 16': ['. Presupuesto por Componente', 14],
                'Formato 17': ['. Fuentes de Financiamiento', 7]
                }

    for formato, lista in titulos_y_columnas.items():
        if formato == 'Formato 11':
            fichas=estructurar_fichas(id)
            for ficha_id, ficha in fichas.items():
                # 2. Crea una hoja nueva (o la sobreescribe si ya existía con el mismo nombre)
                nombre_hoja = f'Formato 11 - {ficha["parte1"]["codigo_ordenado"]}'  # ej. "Actividad 1", "Componente 2", etc.
                ws = wb.create_sheet(title=nombre_hoja)
                crear_encabezado(ws, 'Formato 11. Ficha Técnica  de Indicadores', 4, nombre, fill = True)
                # 3. Aplica el formato base a esa hoja
                formato_esqueleto_fichas(ws)
                # 4. Llama a la función que llena la ficha
                llenar_ficha_en_hoja(ws, ficha)
                ws["B3"] = nombre
                ws["B4"] = entidad
        else:
            formato
            tittle = str(f"{formato}{lista[0] if isinstance(lista, list) else ''}")            
            wb.create_sheet(formato)
            ws = wb[formato]
            if formato == 'Formato 1':
                #### dar formato a la hoja formato 1
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Nombre del programa','Lugar donde se implementó',	'Objetivo',	'Descripción' ,	'Población objetivo',	'Bienes y servicios que entrega',	'Resultados de evaluaciones',	'Vínculo al documento fuente de información']
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                ws.column_dimensions['A'].width = 35
                for num in range(2, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 20
                    
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)

                llenar_f1(ws, id)
            elif formato =='Formato 2':
                #### dar formato a la hoja formato 2
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols1 =['Nombre del programa','Tipo de programa ','','Objetivo','Población objetivo','Bienes y servicios que provee','Cobertura ','','Institución/ Dirección que coordina el programa','Interdependencias  entre los programas','']
                nombre_cols2 = ['','Valor','Otro (Especifique)','','','','Valor','Otro (especifique)','','Valor','Describir interpendencia']
                for i, col in enumerate(nombre_cols1):
                    cell = ws.cell(row = 3, column = i+1, value= col)
                for i, col in enumerate(nombre_cols2):
                    cell = ws.cell(row = 4, column = i+1, value= col)
                for num in range(1, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 20
                merge_filas(ws, 3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                merge_columns_flexible(ws, 4, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1],12,'d')
                formato_encabezado_tablas(ws, 3, lista[1],12,'d')
                ####----- llenado
                llenar_f2(ws, id)
                ####-----
            elif formato == 'Formato 3':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Actor','','Descripción del tipo de relación con el programa','Posición','Influencia']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                    nombre_cols = ['Tipo','Nombre','','','']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 4, column = i+1, value= col)
                merge_filas(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 3, lista[1],12,'d')
                merge_columns_flexible(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1],12,'d')
                ws.row_dimensions[4].height = 36 
                ws.column_dimensions['A'].width = 21    
                ws.column_dimensions['B'].width = 34 
                ws.column_dimensions['C'].width = 58 
                ws.column_dimensions['D'].width = 33 
                ws.column_dimensions['E'].width = 33 
                ####----- llenado
                llenar_f3(ws, id)
                ####-----
            elif formato == 'Formato 4':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Criterio','Descripción del criterio','Justificación de la elección']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1],12,'d')
                ws.column_dimensions['A'].width = 46    
                ws.column_dimensions['B'].width = 46 
                ws.column_dimensions['C'].width = 46 
                ####----- llenado
                llenar_f4(ws, id)
                ####-----
                
            elif formato == 'Árbol de Problemas' or formato == 'Árbol de Objetivos':
                data_map = {
                    'Árbol de Problemas': arbol_problemas,
                    'Árbol de Objetivos': arbol_objetivos
                    }
                #### Llenar Arboles
                crear_encabezado(ws, tittle, lista, nombrePP = nombre, fill = True)
                if tittle in data_map:
                    if data_map[tittle] is not None:
                        arbol = crear_arbol(data_map[tittle])
                        pil_img = Image.open(arbol)
                        output_arbol = io.BytesIO()
                        pil_img.save(output_arbol, format='PNG')
                        output_arbol.seek(0)
                        img = XLImage(output_arbol)
                        ws.add_image(img, "A3")
            elif formato =='Formato 5':
                #### dar formato a la hoja formato 5
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols1 =['Tipo','Descripcion','Hombres','Mujeres','Hablantes de lengua indígena','Grupos de edad','Otros criterios','','Medio de Verificación']
                nombre_cols2 = ['','','','','','','Descripción','Cuantificación','']
                for i, col in enumerate(nombre_cols1):
                    cell = ws.cell(row = 3, column = i+1, value= col)
                for i, col in enumerate(nombre_cols2):
                    cell = ws.cell(row = 4, column = i+1, value= col)
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 30
                    elif num == 2:
                        ws.column_dimensions['B'].width = 40
                    elif num == lista[1]:
                        ws.column_dimensions[chr(num+64)].width = 40
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 17
                if test_formato5:        
                    Llenar_f5(ws, test_formato5)
                merge_filas(ws, 3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1], 12, 'd')
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')

            elif formato =='Formato 6':
                #### dar formato a la hoja formato 6
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                for num in range(1, lista[1]+1):
                    if num == 1:
                        width = 46
                    elif num == 4:
                        width = 24
                    else: 
                        width = 19
                    ws.column_dimensions[chr(num+64)].width = width
                ws.row_dimensions[5].height = 25

                nombre_cols = ['Nombre del municipio*', 'Número de localidades*', 'Población Total *','Población Objetivo **','','',
                               '','','','','','']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col) 
                
                merge_columns_flexible(ws, 3, lista[1], merge_count=8)
                formato_encabezado_tablas(ws, 3, lista[1], 10, 'd')

                nombre_cols2 = ['','','','Cuantificación','Número de habtantes por tamaño de localidad','',
                                '','','','','% de la población urbana*','% de la población rural*']
                for i, col in enumerate(nombre_cols2):
                    ws.cell(row = 4, column = i+1, value= col)
                merge_columns_flexible(ws, 4, lista[1], merge_count=5)
                formato_encabezado_tablas(ws, 4, lista[1], 10, 'd')

                nombre_cols3 = ['','','','','De hasta 500 habitantes','501-2,500',
                                '2,501-10,000','1,001-15,000','15,000-49,999','Más de 50,000','','']
                for i, col in enumerate(nombre_cols3):
                    ws.cell(row = 5, column = i+1, value= col)
                formato_encabezado_tablas(ws, 5, lista[1], 10, 'd')
                nombre_cols4 = ['','número','=+SUMA(C7:C115)','número= suma de las columnas de número de habitantes por tamaño de localidad','número','número','número','número','número','número','Porcentaje = número de habitantes menores o iguales a  2500)/cuentificación','Porcentaje = número de habitantes mayores 2500)/cuentificación']
                for i, col in enumerate(nombre_cols4):
                    cell = ws.cell(row = 6, column = i+1, value= col)
                    formato_celda_cuerpo(cell, 10)

                merge_dos_filas(ws,3, lista[1])
                merge_filas(ws,4, lista[1])
                ##### -------- llenado
                llenar_f6(ws, id)
                ##### -------- 
            elif formato =='Formato 7':
                #### dar formato a la hoja formato 7
                Llenar_f7(ws, id, tittle, lista, nombre)    
            elif formato =='Formato 8':
                #### dar formato a la hoja formato 8
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Bien o servicio',
                            'Descripción del bien o servicio',
                            'Criterios de calidad',
                            'Criterios para determinar la entrega oportuna']
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                ws.column_dimensions['A'].width = 25
                for num in range(2, lista[1]):
                    ws.column_dimensions[chr(num+64)].width = 40
                ws.column_dimensions['D'].width = 50
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)  
                if test_formato8:
                    Llenar_f8(ws, test_formato8)
            elif formato =='Formato 9':
                #### dar formato a la hoja formato 9
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Dependencia /Entidad',
                            'Área',
                            'Código Centro de Costos (3 niveles)',
                            'Función en la ejecución del programa',
                            'Interactúa con',
                            'Mecanismos de coordinación	',
                            'Responsabilidad',
                            'Atribución legal CAPY /RECAPY']
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                for num in range(1, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 22
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)  
                #####------- llenado
                llenar_f9(ws, id)
                #####------- llenado
                    
            elif formato =='Formato 10':
                #### formato 10###
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Resumen Narrativo',
                            '',
                            'Indicadores',
                            'Medio de verificación',
                            'Supuestos']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)  
                nombre_cols2 = ['Tipo','Objetivo','','','']
                for i, col in enumerate(nombre_cols2):
                    ws.cell(row = 4, column = i+1, value= col)
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions[chr(num+64)].width = 20
                    elif num == 2 or num == lista[1]:
                        ws.column_dimensions[chr(num+64)].width = 40
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 22
                merge_filas(ws,3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                merge_columns_flexible(ws, 4, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1], 12, 'd')
                ## llenado de datos
                if test_formato10:
                    llenar_f10(ws, test_formato10)
            elif formato == 'Formato 12':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Nombre del Indicador', 'Descripción de la variable', 'Registro Administrativo',
                                    'Desagregación por sexo', 'Instrumento de recolección de la información',
                                    '¿En qué programa informático/software tiene o tendrá su base de datos?',
                                    'Responsable de la producción de información', 'Periodicidad de la producción de la información']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col) 
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'r')
                ws.row_dimensions[3].height = 65
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 38
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 20
                llenar_f12(ws, id)
            elif formato == 'Formato 13':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Nombre del reporte', 'Descripción general de la información reportada', 'Periodicidad',
                                    'Responsable de la integración']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1], 10, 'r')
                ws.row_dimensions[3].height = 46 
                for num in range(1, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 38
                ####----- llenado
                llenar_f13(ws, id)
                ####-----
            elif formato == 'Formato 14':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                formato_encabezado_tablas(ws, 3, 1, 11, 'd')
                nombre_cols0 = ['Marco de Resultados de mediano plazo', '', '','','','','','','','']
                
                for i, col in enumerate(nombre_cols0):
                    ws.cell(row = 3, column = i+1, value= col)
                nombre_cols = ['Resumen narrativo', 'Indicadores y metas', '','','','','','','',
                            'Medios de verificación']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 4, column = i+1, value= col) 
                nombre_cols2 = ['','Indicadores','Línea base','Metas por año','','','','','','']
                for i, col in enumerate(nombre_cols2):
                    ws.cell(row = 5, column = i+1, value= col)
                nombre_cols3 = ['','','',2025,2026,2027,2028,2029,2030,'']
                for i, col in enumerate(nombre_cols3):
                    ws.cell(row = 6, column = i+1, value= col)
                merge_dos_filas(ws,4, lista[1])
                merge_filas(ws,5, lista[1])
                merge_columns_flexible(ws, 4, lista[1], merge_count=7)
                merge_columns_flexible(ws, 5, lista[1], merge_count=5)
                merge_columns_flexible(ws, 3, lista[1], merge_count=9)
                for fila in range(4, 7):  # 4, 5, 6
                    formato_encabezado_tablas(ws, fila, lista[1],12,'r')
                ws.column_dimensions['A'].width = 55
                ws.column_dimensions['B'].width = 13
                ws.column_dimensions['C'].width = 17
                for col in ['D', 'E', 'F', 'G', 'H', 'I']:
                    ws.column_dimensions[col].width = 10
                ws.column_dimensions['J'].width = 16
                llenar_f14(ws, id_pp)
            elif formato == 'Formato 15':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'r')
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'r')
                nombre_cols = ['Concepto', 'Total de la población objetivo', 'Población programada a atender','','','','','','']
                for i, col in enumerate(nombre_cols):
                    cell = ws.cell(row = 3, column = i+1, value= col) 
                nombre_cols2 = ['', '', 2025,2026,2027,2028,2029,2030]
                for i, col in enumerate(nombre_cols2):
                    cell = ws.cell(row = 4, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1], 12,'r')
                formato_encabezado_tablas(ws, 4, lista[1], 12,'r')
                merge_columns_flexible(ws, 3, lista[1], merge_count=4)
                merge_filas(ws,3, lista[1])
                merge_filas(ws,4, lista[1])
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 40
                    if num == 2:
                        ws.column_dimensions['B'].width = 32 
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 11
                llenar_f15(ws, id)
                ####----- llenado
                ####-----
            elif formato == 'Formato 16':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill=True)
                nombre_cols = ['Componente/ capítulo','Meta de mediano plazo del componente',2025, 'Presupuesto',2026, 'Presupuesto',
                                2027, 'Presupuesto',2028, 'Presupuesto',2029, 'Presupuesto',2030, 'Presupuesto']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col) 
                formato_encabezado_tablas(ws, 3, lista[1],10,'r')
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 38
                    if num == 2:
                        ws.column_dimensions['B'].width = 25
                    else:
                        ws.column_dimensions[chr(num+64)].width = 16 

                    llenar_f16(ws, id)
            else:
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Fuentes de financiamiento',2025, 2026, 2027, 2028, 2029, 2030]
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1], 10, 'd')
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 40
                    else:
                        ws.column_dimensions['B'].width = 10
                llenar_f17(ws, id)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    return wb

def generar_excel(request, id_pp):
    programa = programas_p.objects.get(id_pp=id_pp)
    nombre = programa.nombre_pp
    id = programa.id_pp
    entidad = programa.user.userprofile.siglas

    #arboles
    arbol_problemas= {'nodos':[], 'conexiones':[]}
    if Nodo.objects.filter(id_pp__id_pp=id).exists():
            arbol_problemas = estructura_AP(id)
    arbol_objetivos = {'nodos':[], 'conexiones':[]}
    if Nodo.objects.filter(id_pp__id_pp=id).exists() and NodoAO.objects.filter(id_pp__id_pp=id).exists():
            arbol_objetivos = estructura_AO(id)

    test_formato1 = {}

    test_formato2 = {}

    test_formato3 = {}

    test_formato4 = {}

    test_formato5 ={}
    if PoblacionObjetivo.objects.filter(id_pp=id).exists():
        test_formato5 = estructura_poblacion_obj(id)
    
    test_formato6 = {}
        

    test_formato8 ={}
    if BienServicio.objects.filter(id_pp=id).exists():
        bienes_guardados = BienServicio.objects.filter(id_pp=id)
        test_formato8 = [model_to_dict(bien) for bien in bienes_guardados]    

    test_formato9 = {}
    
    test_formato10= {}
    if FichaIndicador.objects.filter(id_pp_id=id,tipo_ficha='fin').exists():
        test_formato10 = generar_data_mir(id)
    
    test_formato11 = {}
    test_formato12 = {}
    test_formato13 = {}
    test_formato14 = {}
    test_formato15 = {}
    test_formato16 = {}
    test_formato17 = {}

    wb = Workbook()
    titulos_y_columnas = { 
                'Formato 1': ['. Documentación de buenas prácticas y programas similares', 8],
                'Formato 2': ['. Vinculación con otros programas implementados en el estado', 11],
                'Formato 3': ['. Identificación de involucrados', 5],
                'Formato 4': ['. Criterios para la focalización de la población objetivo', 3],
                'Árbol de Problemas': 14,
                'Árbol de Objetivos': 14,
                'Formato 5': ['. Identificación y cuantificación de la población objetivo', 9],
                'Formato 6': ['. Cobertura geográfica', 12],
                'Formato 7': ['. Alineación con la planeación del desarrollo', 2],
                'Formato 8': ['. Características de los bienes y/o servicios del programa', 4],
                'Formato 9': ['. Corresponsabilidad  interinstitucional', 8],
                'Formato 10': ['. Matriz de Indicadores para Resultados (MIR)', 5],
                'Formato 11': ['. Ficha de indicadores', 4],
                'Formato 12': ['. Fuentes de Información', 8], 
                'Formato 13': ['. Informes de desempeño', 4],
                'Formato 14': ['. Marco de resultados en el mediano plazo', 10], 
                'Formato 15': ['. Programación de la atención a la población objetivo en el mediano plazo', 8],
                'Formato 16': ['. Presupuesto por Componente', 14],
                'Formato 17': ['. Fuentes de Financiamiento', 7]
                }

    for formato, lista in titulos_y_columnas.items():
        if formato == 'Formato 11':
            fichas=estructurar_fichas(id)
            for ficha_id, ficha in fichas.items():
                # 2. Crea una hoja nueva (o la sobreescribe si ya existía con el mismo nombre)
                nombre_hoja = f'Formato 11 - {ficha["parte1"]["codigo_ordenado"]}'  # ej. "Actividad 1", "Componente 2", etc.
                ws = wb.create_sheet(title=nombre_hoja)
                crear_encabezado(ws, 'Formato 11. Ficha Técnica  de Indicadores', 4, nombre, fill = True)
                # 3. Aplica el formato base a esa hoja
                formato_esqueleto_fichas(ws)
                # 4. Llama a la función que llena la ficha
                llenar_ficha_en_hoja(ws, ficha)
                ws["B3"] = nombre
                ws["B4"] = entidad
        else:
            formato
            tittle = str(f"{formato}{lista[0] if isinstance(lista, list) else ''}")            
            wb.create_sheet(formato)
            ws = wb[formato]
            if formato == 'Formato 1':
                #### dar formato a la hoja formato 1
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Nombre del programa','Lugar donde se implementó',	'Objetivo',	'Descripción' ,	'Población objetivo',	'Bienes y servicios que entrega',	'Resultados de evaluaciones',	'Vínculo al documento fuente de información']
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                ws.column_dimensions['A'].width = 35
                for num in range(2, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 20
                    
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)

                llenar_f1(ws, id)
            elif formato =='Formato 2':
                #### dar formato a la hoja formato 2
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols1 =['Nombre del programa','Tipo de programa ','','Objetivo','Población objetivo','Bienes y servicios que provee','Cobertura ','','Institución/ Dirección que coordina el programa','Interdependencias  entre los programas','']
                nombre_cols2 = ['','Valor','Otro (Especifique)','','','','Valor','Otro (especifique)','','Valor','Describir interpendencia']
                for i, col in enumerate(nombre_cols1):
                    cell = ws.cell(row = 3, column = i+1, value= col)
                for i, col in enumerate(nombre_cols2):
                    cell = ws.cell(row = 4, column = i+1, value= col)
                for num in range(1, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 20
                merge_filas(ws, 3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                merge_columns_flexible(ws, 4, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1],12,'d')
                formato_encabezado_tablas(ws, 3, lista[1],12,'d')
                ####----- llenado
                llenar_f2(ws, id)
                ####-----
            elif formato == 'Formato 3':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Actor','','Descripción del tipo de relación con el programa','Posición','Influencia']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                    nombre_cols = ['Tipo','Nombre','','','']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 4, column = i+1, value= col)
                merge_filas(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 3, lista[1],12,'d')
                merge_columns_flexible(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1],12,'d')
                ws.row_dimensions[4].height = 36 
                ws.column_dimensions['A'].width = 21    
                ws.column_dimensions['B'].width = 34 
                ws.column_dimensions['C'].width = 58 
                ws.column_dimensions['D'].width = 33 
                ws.column_dimensions['E'].width = 33 
                ####----- llenado
                llenar_f3(ws, id)
                ####-----
            elif formato == 'Formato 4':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Criterio','Descripción del criterio','Justificación de la elección']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1],12,'d')
                ws.column_dimensions['A'].width = 46    
                ws.column_dimensions['B'].width = 46 
                ws.column_dimensions['C'].width = 46 
                ####----- llenado
                llenar_f4(ws, id)
                ####-----
                
            elif formato == 'Árbol de Problemas' or formato == 'Árbol de Objetivos':
                data_map = {
                    'Árbol de Problemas': arbol_problemas,
                    'Árbol de Objetivos': arbol_objetivos
                    }
                #### Llenar Arboles
                crear_encabezado(ws, tittle, lista, nombrePP = nombre, fill = True)
                if tittle in data_map:
                    if data_map[tittle] is not None:
                        arbol = crear_arbol(data_map[tittle])
                        pil_img = Image.open(arbol)
                        output_arbol = io.BytesIO()
                        pil_img.save(output_arbol, format='PNG')
                        output_arbol.seek(0)
                        img = XLImage(output_arbol)
                        ws.add_image(img, "A3")
            elif formato =='Formato 5':
                #### dar formato a la hoja formato 5
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols1 =['Tipo','Descripcion','Hombres','Mujeres','Hablantes de lengua indígena','Grupos de edad','Otros criterios','','Medio de Verificación']
                nombre_cols2 = ['','','','','','','Descripción','Cuantificación','']
                for i, col in enumerate(nombre_cols1):
                    cell = ws.cell(row = 3, column = i+1, value= col)
                for i, col in enumerate(nombre_cols2):
                    cell = ws.cell(row = 4, column = i+1, value= col)
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 30
                    elif num == 2:
                        ws.column_dimensions['B'].width = 40
                    elif num == lista[1]:
                        ws.column_dimensions[chr(num+64)].width = 40
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 17
                if test_formato5:        
                    Llenar_f5(ws, test_formato5)
                merge_filas(ws, 3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1], 12, 'd')
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')

            elif formato =='Formato 6':
                #### dar formato a la hoja formato 6
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                for num in range(1, lista[1]+1):
                    if num == 1:
                        width = 46
                    elif num == 4:
                        width = 24
                    else: 
                        width = 19
                    ws.column_dimensions[chr(num+64)].width = width
                ws.row_dimensions[5].height = 25

                nombre_cols = ['Nombre del municipio*', 'Número de localidades*', 'Población Total *','Población Objetivo **','','',
                               '','','','','','']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col) 
                
                merge_columns_flexible(ws, 3, lista[1], merge_count=8)
                formato_encabezado_tablas(ws, 3, lista[1], 10, 'd')

                nombre_cols2 = ['','','','Cuantificación','Número de habtantes por tamaño de localidad','',
                                '','','','','% de la población urbana*','% de la población rural*']
                for i, col in enumerate(nombre_cols2):
                    ws.cell(row = 4, column = i+1, value= col)
                merge_columns_flexible(ws, 4, lista[1], merge_count=5)
                formato_encabezado_tablas(ws, 4, lista[1], 10, 'd')

                nombre_cols3 = ['','','','','De hasta 500 habitantes','501-2,500',
                                '2,501-10,000','1,001-15,000','15,000-49,999','Más de 50,000','','']
                for i, col in enumerate(nombre_cols3):
                    ws.cell(row = 5, column = i+1, value= col)
                formato_encabezado_tablas(ws, 5, lista[1], 10, 'd')
                nombre_cols4 = ['','número','=+SUMA(C7:C115)','número= suma de las columnas de número de habitantes por tamaño de localidad','número','número','número','número','número','número','Porcentaje = número de habitantes menores o iguales a  2500)/cuentificación','Porcentaje = número de habitantes mayores 2500)/cuentificación']
                for i, col in enumerate(nombre_cols4):
                    cell = ws.cell(row = 6, column = i+1, value= col)
                    formato_celda_cuerpo(cell, 10)

                merge_dos_filas(ws,3, lista[1])
                merge_filas(ws,4, lista[1])
                ##### -------- llenado
                llenar_f6(ws, id)
                ##### -------- 
            elif formato =='Formato 7':
                #### dar formato a la hoja formato 7
                Llenar_f7(ws, id, tittle, lista, nombre)    
            elif formato =='Formato 8':
                #### dar formato a la hoja formato 8
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Bien o servicio',
                            'Descripción del bien o servicio',
                            'Criterios de calidad',
                            'Criterios para determinar la entrega oportuna']
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                ws.column_dimensions['A'].width = 25
                for num in range(2, lista[1]):
                    ws.column_dimensions[chr(num+64)].width = 40
                ws.column_dimensions['D'].width = 50
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)  
                if test_formato8:
                    Llenar_f8(ws, test_formato8)
            elif formato =='Formato 9':
                #### dar formato a la hoja formato 9
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Dependencia /Entidad',
                            'Área',
                            'Código Centro de Costos (3 niveles)',
                            'Función en la ejecución del programa',
                            'Interactúa con',
                            'Mecanismos de coordinación	',
                            'Responsabilidad',
                            'Atribución legal CAPY /RECAPY']
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                for num in range(1, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 22
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)  
                #####------- llenado
                llenar_f9(ws, id)
                #####------- llenado
                    
            elif formato =='Formato 10':
                #### formato 10###
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Resumen Narrativo',
                            '',
                            'Indicadores',
                            'Medio de verificación',
                            'Supuestos']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)  
                nombre_cols2 = ['Tipo','Objetivo','','','']
                for i, col in enumerate(nombre_cols2):
                    ws.cell(row = 4, column = i+1, value= col)
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions[chr(num+64)].width = 20
                    elif num == 2 or num == lista[1]:
                        ws.column_dimensions[chr(num+64)].width = 40
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 22
                merge_filas(ws,3, lista[1])
                merge_columns_flexible(ws, 3, lista[1])
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'd')
                merge_columns_flexible(ws, 4, lista[1])
                formato_encabezado_tablas(ws, 4, lista[1], 12, 'd')
                ## llenado de datos
                if test_formato10:
                    llenar_f10(ws, test_formato10)
            elif formato == 'Formato 12':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Nombre del Indicador', 'Descripción de la variable', 'Registro Administrativo',
                                    'Desagregación por sexo', 'Instrumento de recolección de la información',
                                    '¿En qué programa informático/software tiene o tendrá su base de datos?',
                                    'Responsable de la producción de información', 'Periodicidad de la producción de la información']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col) 
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'r')
                ws.row_dimensions[3].height = 65
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 38
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 20
                llenar_f12(ws, id)
            elif formato == 'Formato 13':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Nombre del reporte', 'Descripción general de la información reportada', 'Periodicidad',
                                    'Responsable de la integración']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1], 10, 'r')
                ws.row_dimensions[3].height = 46 
                for num in range(1, lista[1]+1):
                    ws.column_dimensions[chr(num+64)].width = 38
                ####----- llenado
                llenar_f13(ws, id)
                ####-----
            elif formato == 'Formato 14':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                formato_encabezado_tablas(ws, 3, 1, 11, 'd')
                nombre_cols0 = ['Marco de Resultados de mediano plazo', '', '','','','','','','','']
                
                for i, col in enumerate(nombre_cols0):
                    ws.cell(row = 3, column = i+1, value= col)
                nombre_cols = ['Resumen narrativo', 'Indicadores y metas', '','','','','','','',
                            'Medios de verificación']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 4, column = i+1, value= col) 
                nombre_cols2 = ['','Indicadores','Línea base','Metas por año','','','','','','']
                for i, col in enumerate(nombre_cols2):
                    ws.cell(row = 5, column = i+1, value= col)
                nombre_cols3 = ['','','',2025,2026,2027,2028,2029,2030,'']
                for i, col in enumerate(nombre_cols3):
                    ws.cell(row = 6, column = i+1, value= col)
                merge_dos_filas(ws,4, lista[1])
                merge_filas(ws,5, lista[1])
                merge_columns_flexible(ws, 4, lista[1], merge_count=7)
                merge_columns_flexible(ws, 5, lista[1], merge_count=5)
                merge_columns_flexible(ws, 3, lista[1], merge_count=9)
                for fila in range(4, 7):  # 4, 5, 6
                    formato_encabezado_tablas(ws, fila, lista[1],12,'r')
                ws.column_dimensions['A'].width = 55
                ws.column_dimensions['B'].width = 13
                ws.column_dimensions['C'].width = 17
                for col in ['D', 'E', 'F', 'G', 'H', 'I']:
                    ws.column_dimensions[col].width = 10
                ws.column_dimensions['J'].width = 16
                llenar_f14(ws, id_pp)
            elif formato == 'Formato 15':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'r')
                formato_encabezado_tablas(ws, 3, lista[1], 12, 'r')
                nombre_cols = ['Concepto', 'Total de la población objetivo', 'Población programada a atender','','','','','','']
                for i, col in enumerate(nombre_cols):
                    cell = ws.cell(row = 3, column = i+1, value= col) 
                nombre_cols2 = ['', '', 2025,2026,2027,2028,2029,2030]
                for i, col in enumerate(nombre_cols2):
                    cell = ws.cell(row = 4, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1], 12,'r')
                formato_encabezado_tablas(ws, 4, lista[1], 12,'r')
                merge_columns_flexible(ws, 3, lista[1], merge_count=4)
                merge_filas(ws,3, lista[1])
                merge_filas(ws,4, lista[1])
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 40
                    if num == 2:
                        ws.column_dimensions['B'].width = 32 
                    else: 
                        ws.column_dimensions[chr(num+64)].width = 11
                llenar_f15(ws, id)
                ####----- llenado
                ####-----
            elif formato == 'Formato 16':
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill=True)
                nombre_cols = ['Componente/ capítulo','Meta de mediano plazo del componente',2025, 'Presupuesto',2026, 'Presupuesto',
                                2027, 'Presupuesto',2028, 'Presupuesto',2029, 'Presupuesto',2030, 'Presupuesto']
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col) 
                formato_encabezado_tablas(ws, 3, lista[1],10,'r')
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 38
                    if num == 2:
                        ws.column_dimensions['B'].width = 25
                    else:
                        ws.column_dimensions[chr(num+64)].width = 16 

                    llenar_f16(ws, id)
            else:
                crear_encabezado(ws, tittle, lista[1], nombrePP = nombre, fill = True)
                nombre_cols = ['Fuentes de financiamiento',2025, 2026, 2027, 2028, 2029, 2030]
                for i, col in enumerate(nombre_cols):
                    ws.cell(row = 3, column = i+1, value= col)
                formato_encabezado_tablas(ws, 3, lista[1], 10, 'd')
                for num in range(1, lista[1]+1):
                    if num == 1:
                        ws.column_dimensions['A'].width = 40
                    else:
                        ws.column_dimensions['B'].width = 10
                llenar_f17(ws, id)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    nombre_excel = f"{nombre}.xlsx"
    # Guardar el archivo Excel en un objeto de respuesta HTTP para descargar
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_excel}"'
    output_file = f"excel_{id}"
    wb.save(response)
    return response