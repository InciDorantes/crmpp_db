from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.worksheet.merge import MergedCell
import os
from django.conf import settings
import io
from io import BytesIO
import graphviz
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
os.environ["PATH"] += os.pathsep + 'C:\\Program Files (x86)\\Graphviz\\bin'
from graphviz import Digraph
from openpyxl.styles import Alignment
from PPRegister.models import FormatoUno, FormatoDos, FormatoTres, FormatoCuatro, Formato9,UserProfile, Formato6, Municipios, APED, Directriz, Vertiente, ObjetivoEstrategico, ObjetivoEspecifico, LineaAccion, FormatoCatorce, FormatoDieciseis, FormatoDoce, FormatoQuince, FormatoTrece, LineaBase
from openpyxl.utils import get_column_letter
##### arboles #####
def wrap_text(text, words_per_line):
    words = text.split()  # Split the text into a list of words
    lines = [' '.join(words[i:i+words_per_line]) for i in range(0, len(words), words_per_line)]
    return '\n'.join(lines)

def detectar_tipo_arbol(data):
    if 'problema_central' in data:
        return 'problema'
    else:
        return 'solucion'


def crear_arbol(data):
    dot = Digraph(format='png', graph_attr={'bgcolor': 'transparent', 'splines': 'ortho'})
    dot.attr(rankdir='BT')  # Flechas de abajo hacia arriba

    tipo_labels = {
        1: "ci",  # causa indirecta
        2: "cd",  # causa directa
        3: "pc",  # problema central
        4: "ed",  # efecto directo
        5: "ei",  # efecto indirecto
        6: "es",  # efecto superior
    }

    nodos_dict = {}

    # Crear nodos
    for nodo in data["nodos"]:
        tipo = nodo["tipo"]
        id_nodo = nodo["id_nodo"]
        texto = nodo["texto"]
        key = f"{tipo_labels.get(tipo, 'x')}_{id_nodo}"
        nodos_dict[id_nodo] = key  # guardar mapeo para usar en conexiones
        dot.node(key, wrap_text(texto, 10 if tipo in {3,6} else 3), shape='rectangle',
                 fontname='Lato Semibold', fontsize='10')

    # Crear conexiones
    for conexion in data["conexiones"]:
        origen_id = conexion["origen_id"]
        destino_id = conexion["destino_id"]
        if origen_id in nodos_dict and destino_id in nodos_dict:
            dot.edge(nodos_dict[origen_id], nodos_dict[destino_id])

     # Renderizar en memoria
    img_bytes = dot.pipe(format='png')
    return BytesIO(img_bytes)


##### Estilos ######
# dar formato a encabezado
def crear_encabezado(ws, titulo, columnas, nombrePP, fill, pintar_programa=True):
    from openpyxl.styles import PatternFill, Alignment, Font
    import os
    from openpyxl.drawing.image import Image
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columnas)
    cell = ws.cell(row=1, column=1, value=titulo)
    cell.alignment = Alignment(vertical='center')
    cell.font = Font(bold=True, size=18, name='Lato')
    ws.row_dimensions[1].height = 50

    if pintar_programa:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columnas)
        cell = ws.cell(row=2, column=1, value=f"Programa Presupuestario: {nombrePP}")
        cell.font = Font(bold=True, size=12, name='Lato', color='FFFFFF' if fill else '000000')
        if fill:
            cell.fill = PatternFill(start_color="970E48", end_color="970E48", fill_type="solid")

    # add logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'logoyuc.png')

    if columnas > 5:
        logo_column = columnas - 1
    elif columnas < 3:
        logo_column = 3
    else:
        logo_column = columnas
    img_logo = Image(logo_path)
    img_logo.width = 307
    img_logo.height = 50

    # Si pintar_programa es False, el logo debería ir en la fila 1, si es True, podría ir en fila 1 o 2 según diseño.
    # Aquí lo dejamos en fila 1 para simplificar:
    ws.add_image(img_logo, f"{chr(64 + logo_column)}1")


def formato_encabezado_tablas(ws, fila, columna, num_fuente, color):
    # dar formato a celdas
    if color == 'd':
        color = "c2995c"
        font_color = '000000'
    else:
        color = "970e48"
        font_color = "FFFFFF"

    for i in range(1, columna+1):
        cell = ws.cell(row=fila, column=i)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True, size=num_fuente, name= 'Lato', color= font_color)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def formato_celda_cuerpo(cell, num):
    cell.alignment = Alignment(wrap_text=True, vertical= 'center')
    cell.font = Font(size=num, name= 'Lato')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def formato_esqueleto_fichas(ws):
    secciones = {'parte1': ['Programa Presupuestario','Dependencia/ entidad','Nivel del objetivo','Objetivo','Nombre del indicador'], 
                 'Meta datos del Indicador': ['Definición','Tipo de Algoritmo','Tipo de indicador','Periodicidad de cálculo','Tendencia','Ámbito de medición','Dimensión del desempeño'], 
                 'Variables': ['Variables','Nombre','','Medio de verificación'], 
                 'Línea base o valor de referencia': ['Variables','Valor', 'Unidad de Medida','Fecha (día, mes y año)'] , 
                 'Meta': ['Variables','Valor', 'Unidad de Medida','Fecha (día, mes y año)']}
    
    def fmr_celdas_fichas(cell, color, bold, italic):
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.font = Font(size=10, name='Lato', bold=bold, italic=italic)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')      
        
    def fmt_seccion(cell):
        cell.font = Font(size=10, name='Lato', bold=True,color = 'ffffff')
        cell.fill = PatternFill(start_color='970e47', end_color='970e47', fill_type='solid')
        cell.alignment = Alignment(horizontal = 'center', vertical= 'center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.column_dimensions['A'].width = 50.29
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 35.86
    for i, (seccion, elemento) in enumerate(secciones.items()):
        if seccion == 'parte1':
            for idx, row in enumerate(elemento):
                cell = ws.cell(row=3 + idx, column=1, value=row)
                fmr_celdas_fichas(cell,'c2995c', True, False )
                ws.merge_cells(start_row = 3 + idx, start_column = 2, end_row = 3 + idx, end_column = 4)
                for col in range(2, 5):
                    cell_merged = ws.cell(row=3 + idx, column=col)
                    formato_celda_cuerpo(cell_merged, 10)
        elif seccion == 'Meta datos del Indicador':
            ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=4)
            cell = ws.cell (row = 8, column = 1, value = seccion)
            fmt_seccion(cell)
            for idx, row in enumerate(elemento):
                cell = ws.cell(row=9 + idx, column=1, value=row)
                fmr_celdas_fichas(cell, 'c2995c', False, True)
                ws.merge_cells(start_row=9 + idx, start_column=2, end_row=9 + idx, end_column=4)
                # Apply formatting to all merged cells in columns 2-4
                for col in range(2, 5):
                    cell_merged = ws.cell(row=9 + idx, column=col)
                    formato_celda_cuerpo(cell_merged, 10)
        elif seccion == 'Variables':
            ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=4)
            cell = ws.cell (row = 16, column = 1, value = seccion)
            fmt_seccion(cell)
            for idx, col in enumerate(elemento):
                cell = ws.cell (row = 17, column = idx+1, value = col)
                fmr_celdas_fichas(cell, 'd9d9d9', False, True)
                merge_columns_flexible(ws,17,4, merge_count=1)
                cell.alignment = Alignment(horizontal = 'center', vertical= 'center')
            for idx, row in enumerate(['B', 'C', 'D', 'E']):
                cell = ws.cell(row=18 + idx, column=1, value=f'Variable {row}')
                fmr_celdas_fichas(cell,'ffffff', False, False )
                # Merge columns 2 and 3 in each row
                ws.merge_cells(start_row=18 + idx, start_column=2, end_row=18 + idx, end_column=3)
                # Add formato_celda_cuerpo to empty cells in columns 2-4
                for col in range(2, 5):
                    cell_empty = ws.cell(row=18 + idx, column=col)
                    if cell_empty.value is None or cell_empty.value == '':
                        formato_celda_cuerpo(cell_empty, 10)

        elif seccion == 'Línea base o valor de referencia':
            ws.merge_cells(start_row=22, start_column=1, end_row=22, end_column=4)
            cell = ws.cell (row = 22, column = 1, value = seccion)
            fmt_seccion(cell)
            for idx, col in enumerate(elemento):
                cell = ws.cell (row = 23, column = idx+1, value = col)
                fmr_celdas_fichas(cell, 'd9d9d9', False, True)
                cell.alignment = Alignment(horizontal = 'center', vertical= 'center')
            for idx, row in enumerate(['B', 'C', 'D', 'E', 'Línea base o valor de referencia']):
                cell = ws.cell(row=24 + idx, column=1, value=f'Variable {row}' if row != 'Línea base o valor de referencia' else row )
                fmr_celdas_fichas(cell,'ffffff', False, False )
                for col in range(2, 5):
                        cell_empty = ws.cell(row=24 + idx, column=col)
                        if cell_empty.value is None or cell_empty.value == '':
                            formato_celda_cuerpo(cell_empty, 10)
        elif seccion == 'Meta':
            ws.merge_cells(start_row=29, start_column=1, end_row=29, end_column=4)
            cell = ws.cell (row = 29, column = 1, value = seccion) 
            fmt_seccion(cell)
            for idx, col in enumerate(elemento):
                cell = ws.cell (row = 30, column = idx+1, value = col)
                cell.alignment = Alignment(horizontal = 'center', vertical= 'center')
                fmr_celdas_fichas(cell, 'd9d9d9', False, True)
            for idx, row in enumerate(['B', 'C', 'D', 'E', 'Meta']):
                cell = ws.cell(row=31 + idx, column=1, value=f'Variable {row}' if row != 'Meta' else row )
                fmr_celdas_fichas(cell,'ffffff', False, False )
                for col in range(2, 5):
                        cell_empty = ws.cell(row=31 + idx, column=col)
                        if cell_empty.value is None or cell_empty.value == '':
                            formato_celda_cuerpo(cell_empty, 10)

def merge_filas(ws, fila, total_columnas):
    """
    For a given row, go through all columns and merge with the cell below if it is empty.
    """
    for col in range(1, total_columnas + 1):
        if ws.cell(row=fila+1, column=col).value == '':
            ws.merge_cells(start_row=fila, start_column=col, end_row=fila+1, end_column=col)

def merge_dos_filas(ws, fila, total_columnas):
    """
    Fusiona verticalmente tres celdas por columna: la de la fila dada y las dos siguientes,
    pero solo si las celdas de abajo están vacías.
    """
    for col in range(1, total_columnas + 1):
        if (ws.cell(row=fila+1, column=col).value == '' and
            ws.cell(row=fila+2, column=col).value == ''):
            ws.merge_cells(start_row=fila, start_column=col, end_row=fila+2, end_column=col)

def merge_columns_flexible(ws, fila, total_columnas, merge_count=1):
    for col in range(1, total_columnas - (merge_count - 1) + 1):  # Adjusted loop condition
        # Verifica que las siguientes 'merge_count' celdas estén vacías
        if all(ws.cell(row=fila, column=col+i).value == '' for i in range(1, merge_count + 1)):
            ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col+merge_count)

   
#################
#### Transformación de datos ####
def crear_lista_plana(value):
    if isinstance(value, list):
        flat = []
        for item in value:
            if isinstance(item, list):
                flat.extend(crear_lista_plana(item))
            else:
                flat.append(item)
        return flat
    else:
        return [value]


# Formato 1
def crear_lista_plana(value):
    if isinstance(value, list):
        flat = []
        for item in value:
            if isinstance(item, list):
                flat.extend(crear_lista_plana(item))
            else:
                flat.append(item)
        return flat
    else:
        return [value]
    
def llenar_f1(ws, id_pp):
    registros = FormatoUno.objects.filter(id_pp=id_pp)
    fila_inicio = 4
    if not registros:
        return
    for i, reg in enumerate(registros, start=fila_inicio):

        formato_celda_cuerpo(ws.cell(row=i, column=1, value=reg.nombre),11)      
        formato_celda_cuerpo(ws.cell(row=i, column=2, value=reg.lugar_implementacion),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=3, value=reg.objetivo),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=4, value=reg.descripcion),11)           
        formato_celda_cuerpo(ws.cell(row=i, column=5, value=reg.poblacion_objetivo),11)     
        formato_celda_cuerpo(ws.cell(row=i, column=6, value=reg.bienes_servicios),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=7, value=reg.resultados_evaluacion),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=8, value=reg.vinculo_documento),11) 

def llenar_f2(ws, id_pp):
    registros = FormatoDos.objects.filter(id_pp = id_pp)
    fila_inicio = 5
    if not registros:
        return
    for i, reg in enumerate(registros, start=fila_inicio):
        formato_celda_cuerpo(ws.cell(row=i, column=1, value=reg.nombre),11)
        if reg.tipo_programa in {'federal', 'estatal', 'municipal'}:
            formato_celda_cuerpo(ws.cell(row=i, column=2, value=str(reg.tipo_programa).capitalize()), 11)
            formato_celda_cuerpo(ws.cell(row=i, column=3, value=''), 11)
        else:
            formato_celda_cuerpo(ws.cell(row=i, column= 2 , value= ''),11)
            formato_celda_cuerpo(ws.cell(row=i, column= 3 , value= str(reg.tipo_programa).capitalize()),11)
        formato_celda_cuerpo(ws.cell(row=i, column=4, value=reg.objetivo),11)
        formato_celda_cuerpo(ws.cell(row=i, column=5, value=reg.poblacion_objetivo),11)
        formato_celda_cuerpo(ws.cell(row=i, column=6, value=reg.bienes_servicios),11)
        
        if reg.cobertura in {'todos', 'regional', 'municipal'}:
            formato_celda_cuerpo(ws.cell(row=i, column=7 , value=str(reg.cobertura).capitalize()),11)
            formato_celda_cuerpo(ws.cell(row=i, column=8 , value=''),11)
        else:
            formato_celda_cuerpo(ws.cell(row=i, column=7 , value=''),11)
            formato_celda_cuerpo(ws.cell(row=i, column=8 , value=str(reg.cobertura).capitalize()),11)
        formato_celda_cuerpo(ws.cell(row=i, column=9, value=reg.institucion_coordinadora),11)
        formato_celda_cuerpo(ws.cell(row=i, column=10, value=reg.interdependencia),11)
        formato_celda_cuerpo(ws.cell(row=i, column=11, value=reg.descripcion_interdependencia),11)



def llenar_f3(ws, id_pp):
    registros = FormatoTres.objects.filter(id_pp=id_pp)
    fila_inicio = 5
    if not registros:
        return
    for i, reg in enumerate(registros, start=fila_inicio):

        formato_celda_cuerpo(ws.cell(row=i, column=1, value=reg.tipo_actor),11)      
        formato_celda_cuerpo(ws.cell(row=i, column=2, value=reg.nombre),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=3, value=reg.descripcion),11)           
        formato_celda_cuerpo(ws.cell(row=i, column=4, value=reg.posicion),11)     
        formato_celda_cuerpo(ws.cell(row=i, column=5, value=reg.influencia),11) 

def llenar_f4(ws, id_pp):
    registros = FormatoCuatro.objects.filter(id_pp=id_pp)
    criterios = [
        'Nivel de ingreso',
        'Sexo',
        'Grupo etario',
        'Condición de hablante de lengua indígena',
        'Ubicación geográfica',
        'Otros criterios (especificar):'
    ]
    fila_inicio = 4

    if not registros:
        return
    
    for i, reg in enumerate(registros, start=fila_inicio):
        valores = [
            (reg.descripcion_nivel_ingreso, reg.justificacion_nivel_ingreso),
            (reg.descripcion_sexo, reg.justificacion_sexo),
            (reg.descripcion_grupo_etario, reg.justificacion_grupo_etario),
            (reg.descripcion_lengua_indigena, reg.justificacion_lengua_indigena),
            (reg.descripcion_ubicacion_geografica, reg.justificacion_ubicacion_geografica),
            (reg.descripcion_otros_criterios, reg.justificacion_otros_criterios)
        ]
        for j, criterio in enumerate(criterios):
            descripcion, justificacion = valores[j]
            formato_celda_cuerpo(ws.cell(row=i + j, column=1, value=criterio), 11)
            formato_celda_cuerpo(ws.cell(row=i + j, column=2, value=descripcion), 11)
            formato_celda_cuerpo(ws.cell(row=i + j, column=3, value=justificacion), 11)
        break

def Llenar_f5(ws, dic):
    if len(dic) == 0:
        return
    for i, (key, val) in enumerate(dic.items()):
        # llenar columna A con el nombre de la key
        if key == 'referencia':
            c = ws.cell(row = i+5, column = 1, value = 'Población de referencia')
            formato_celda_cuerpo(c, 11)
        elif key == 'potencial':
            c = ws.cell(row = i+5, column = 1, value = 'Población potencial o Afectada')
            formato_celda_cuerpo(c, 11)
        elif key == 'objetivo':
            c = ws.cell(row = i+5, column = 1, value = 'Población Objetivo')
            formato_celda_cuerpo(c, 11)
        elif key == 'postergada':
            c= ws.cell(row = i+5, column = 1, value = 'Población Postergada')
            formato_celda_cuerpo(c, 11)
        for col, v in enumerate(val.values()):
            cell = ws.cell(row = i+5, column= col+2, value = v if v else '')
            formato_celda_cuerpo(cell, 11)

def llenar_f6(ws, id_pp):
    registros = Formato6.objects.filter(id_pp=id_pp)
    fila_inicio = 7
    if not registros:
        return
    for i, reg in enumerate(registros, start=fila_inicio):
        municipios = Municipios.objects.filter(id_municipio=reg.id_municipio_id).first()
        municipio = municipios.municipio if municipios else ''
        poblacion = municipios.poblacion if municipios else ''

        formato_celda_cuerpo(ws.cell(row=i, column=1, value=municipio),11)      
        formato_celda_cuerpo(ws.cell(row=i, column=2, value=reg.localidades),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=3, value=poblacion),11)  # <-- aquí cambio
        formato_celda_cuerpo(ws.cell(row=i, column=4, value=reg.cuantificacion_general),11)           
        formato_celda_cuerpo(ws.cell(row=i, column=5, value=reg.tipo1),11)     
        formato_celda_cuerpo(ws.cell(row=i, column=6, value=reg.tipo2),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=7, value=reg.tipo3),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=8, value=reg.tipo4),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=9, value=reg.tipo5),11)
        formato_celda_cuerpo(ws.cell(row=i, column=10, value=reg.tipo6),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=11, value=reg.porcentaje1),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=12, value=reg.porcentaje2),11) 


def Llenar_f7(ws, id_pp):
    row = 4
    salto = 2

    apeds = (
        APED.objects
        .filter(id_pp_id=id_pp)
        .prefetch_related(
            "directriz_set__vertiente_set__objetivoestrategico_set__objetivoespecifico_set"
        )
    )

    for aped in apeds:
        for directriz in aped.directriz_set.all():
            for vertiente in directriz.vertiente_set.all():
                # Directriz y Vertiente (sin merge, siempre simples)
                celda = ws.cell(row=row, column=1, value="Directriz")
                formato_celda_cuerpo(celda, 10)
                celda = ws.cell(row=row, column=2, value=directriz.directriz)
                formato_celda_cuerpo(celda, 10)
                row += 1

                celda = ws.cell(row=row, column=1, value="Vertiente")
                formato_celda_cuerpo(celda, 10)
                celda = ws.cell(row=row, column=2, value=vertiente.vertiente)
                formato_celda_cuerpo(celda, 10)
                row += 1

                # Objetivos estratégicos (sin merge, título en cada fila)
                for objetivo_estrategico in vertiente.objetivoestrategico_set.all():
                    celda = ws.cell(row=row, column=1, value="Objetivo estratégico")
                    formato_celda_cuerpo(celda, 10)
                    celda = ws.cell(row=row, column=2, value=objetivo_estrategico.objetivo_estrategico)
                    formato_celda_cuerpo(celda, 10)
                    row += 1

                # Objetivos específicos: agrupar con merge vertical
                objetivos_especificos = []
                objetivo_especifico_ids = set()

                for objetivo_estrategico in vertiente.objetivoestrategico_set.all():
                    for objetivo_especifico in objetivo_estrategico.objetivoespecifico_set.all():
                        if objetivo_especifico.id_objetivo_especifico not in objetivo_especifico_ids:
                            objetivos_especificos.append(objetivo_especifico)
                            objetivo_especifico_ids.add(objetivo_especifico.id_objetivo_especifico)

                if objetivos_especificos:
                    start_row = row  # fila donde empieza el grupo

                    for oesp in objetivos_especificos:
                        celda = ws.cell(row=row, column=2, value=oesp.objetivo_especifico)
                        formato_celda_cuerpo(celda, 10)
                        row += 1

                    end_row = row - 1  # fila donde termina el grupo

                    # Merge vertical en la columna A para "Objetivo específico"
                    if start_row != end_row:
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    celda_merge = ws.cell(row=start_row, column=1, value="Objetivo específico")
                    formato_celda_cuerpo(celda_merge, 10)

                # Líneas de acción: agrupamos con merge vertical igual
                if objetivos_especificos:
                    lineas_accion = (
                        LineaAccion.objects
                        .filter(id_objetivo_especifico__in=objetivos_especificos)
                        .order_by("linea_accion")
                        .distinct()
                    )
                    lineas_vistas = set()
                    lineas_a_imprimir = []
                    for linea in lineas_accion:
                        if linea.linea_accion not in lineas_vistas:
                            lineas_a_imprimir.append(linea.linea_accion)
                            lineas_vistas.add(linea.linea_accion)

                    if lineas_a_imprimir:
                        start_row = row

                        for linea_acc in lineas_a_imprimir:
                            celda = ws.cell(row=row, column=2, value=linea_acc)
                            formato_celda_cuerpo(celda, 10)
                            row += 1

                        end_row = row - 1
                        if start_row != end_row:
                            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                        celda_merge = ws.cell(row=start_row, column=1, value="Línea de acción")
                        formato_celda_cuerpo(celda_merge, 10)

                row += salto

def Llenar_f8(ws, list):
    start_row = 4  # Assuming row 3 is header, so data starts at row 4
    if len(list) == 0:
        return
    for i, dicc in enumerate(list):
        for key, val in dicc.items():
            if key == 'bien':
                cell = ws.cell(row=start_row + i, column=1, value=val)
                formato_celda_cuerpo(cell, 11)
            elif key == 'descripcion':
                cell = ws.cell(row=start_row + i, column=2, value=val)
                formato_celda_cuerpo(cell, 11)
            elif key == 'criterio_calidad':
                cell = ws.cell(row=start_row + i, column=3, value=val)
                formato_celda_cuerpo(cell, 11)
            elif key == 'criterio':
                cell = ws.cell(row=start_row + i, column=4, value=val)
                formato_celda_cuerpo(cell, 11)

def llenar_f9(ws, id_pp):
    registros = Formato9.objects.filter(id_pp=id_pp)
    fila_inicio = 4
    if not registros:
        return
    for i, reg in enumerate(registros, start=fila_inicio):
        perfil = UserProfile.objects.filter(user=reg.username).first()
        dependencia = perfil.dependencia if perfil else ''
        siglas = perfil.siglas if perfil else ''

        formato_celda_cuerpo(ws.cell(row=i, column=1, value=siglas),11)      
        formato_celda_cuerpo(ws.cell(row=i, column=2, value=dependencia),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=3, value=reg.username.username),11)  # <-- aquí cambio
        formato_celda_cuerpo(ws.cell(row=i, column=4, value=reg.funcion),11)           
        formato_celda_cuerpo(ws.cell(row=i, column=5, value=reg.interactua_con),11)     
        formato_celda_cuerpo(ws.cell(row=i, column=6, value=reg.mecanismo),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=7, value=reg.responsabilidad),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=8, value=reg.atribucion),11) 

def llenar_f10(ws, diccionario):
    if len(diccionario) == 0:
        return
    
    def procesar_mir(mir_data):
        resultados = []
        
        # Nivel Fin
        fin = mir_data["fin"]["fin"]
        resultados.append({
            "Nivel": "Fin",
            "Objetivo": fin["objetivo"],
            "Indicador": fin["indicador"],
            "Medio Verificación": fin["medio_verificacion"] or '',
            "Supuesto": fin["supuestos"] or ''
        })
        
        # Nivel Propósito
        proposito = mir_data["proposito"]["proposito"]
        for idx, indicador in proposito["indicadores"].items():
            resultados.append({
                "Nivel": f"Propósito",
                "Objetivo": proposito["objetivo"],
                "Indicador": indicador["name"],
                "Medio Verificación": indicador["medio"] or '',
                "Supuesto": indicador["supuesto"] or ''
            })
        
        # Nivel Componente y Actividades
        for comp_key, componente in mir_data["componente"].items():
            # Indicadores del Componente
            for idx, indicador in componente["indicadores"].items():
                resultados.append({
                    "Nivel": str(f"{comp_key}").replace('_', ' ').capitalize(),
                    "Objetivo": componente["objetivo"],
                    "Indicador": indicador["nombre"],
                    "Medio Verificación": indicador["medio"] or '',
                    "Supuesto": indicador["supuesto"] or ''
                })
            
            # Actividades del Componente
            for act_key, actividad in componente["actividades"].items():
                for idx, indicador in actividad["indicadores"].items():
                    resultados.append({
                        "Nivel": str(f"{act_key}").replace('_', ' ').capitalize(),
                        "Objetivo": actividad["objetivo"],
                        "Indicador": indicador["nombre"],
                        "Medio Verificación": indicador["medio"] or '',
                        "Supuesto": indicador["supuesto"] or ''
                    })
        
        return pd.DataFrame(resultados)
    
    def merge_celdas_por_nivel(ws, df: pd.DataFrame):
        # Columnas a mergear (Objetivo y Nivel)
        cols_to_merge = ["Nivel", "Objetivo", "Supuesto"]
        
        for col in cols_to_merge:
            col_idx = list(df.columns).index(col) + 1  # +1 porque openpyxl empieza en 1
            current_value = None
            start_row = 2  # Fila inicial (fila 1 es el header)
            
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                
                if cell_value != current_value:
                    if current_value is not None:
                        end_row = row - 1
                        if start_row != end_row:  # Solo mergear si hay múltiples filas
                            ws.merge_cells(
                                start_row=start_row, start_column=col_idx,
                                end_row=end_row, end_column=col_idx
                            )
                            # Centrar texto verticalmente
                            for r in range(start_row, end_row + 1):
                                ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="center")
                                formato_celda_cuerpo(ws.cell(row=r, column=col_idx), 11)
                    
                    current_value = cell_value
                    start_row = row
            # Mergear el último grupo
            if current_value is not None and start_row != ws.max_row:
                ws.merge_cells(
                    start_row=start_row, start_column=col_idx,
                    end_row=ws.max_row, end_column=col_idx
                )
                for r in range(start_row, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="center")
                    formato_celda_cuerpo(ws.cell(row=r, column=col_idx), 11)
    df_mir = procesar_mir(diccionario)
    for r_idx, row in enumerate(dataframe_to_rows(df_mir, index=False, header=False), 5):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            formato_celda_cuerpo(ws.cell(row=r_idx, column=c_idx), 11)
    
    return merge_celdas_por_nivel(ws, df_mir)

def llenar_f13(ws, id_pp):
    registros = FormatoTrece.objects.filter(id_pp=id_pp)
    fila_inicio = 4
    if not registros:
        #print(f"No FormatoTrece records found for id_pp: {id_pp}")  # Debugging
        return  # Exit if no records are found
    for i, reg in enumerate(registros, start=fila_inicio):
        formato_celda_cuerpo(ws.cell(row=i, column=1, value=reg.nombre_reporte),11)
        formato_celda_cuerpo(ws.cell(row=i, column=2, value=reg.descripcion),11)
        formato_celda_cuerpo(ws.cell(row=i, column=3, value=reg.periodicidad),11)
        formato_celda_cuerpo(ws.cell(row=i, column=4, value=reg.responsable_integracion),11)

def llenar_f14(ws, id_pp):
    registros = FormatoCatorce.objects.filter(id_pp=id_pp)
    print('registros:   ')
    print(type(registros))
    fila_inicio = 7
    if not registros:
        return
    for i, reg in enumerate(registros, start=fila_inicio):
        if str(reg.ficha).__contains__('Propósito'):
            val = 'Propósito'
        else:
            val = ''
        formato_celda_cuerpo(ws.cell(row=i, column=1, value= val),11)   
        
        formato_celda_cuerpo(ws.cell(row=i, column=2, value= str(reg.ficha)),11)      
        #print("OBJETO:      ", reg.ficha)
        #print(type(reg.ficha))
        #print("Error 2:     ", reg.variable.id)
        varcita= reg.variable.id
        lineabase = LineaBase.objects.get(variable_id = varcita)
        #print("Error 2:     ",lineabase.resultado_estimado)
        formato_celda_cuerpo(ws.cell(row=i, column=3, value=str(lineabase.resultado_estimado)),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=4, value=reg.meta_2025),11)  # <-- aquí cambio
        formato_celda_cuerpo(ws.cell(row=i, column=5, value=reg.meta_2026),11)           
        formato_celda_cuerpo(ws.cell(row=i, column=6, value=reg.meta_2028),11)     
        formato_celda_cuerpo(ws.cell(row=i, column=7, value=reg.meta_2029),11) 
        formato_celda_cuerpo(ws.cell(row=i, column=8, value=reg.meta_2030),11) 


def llenar_ficha_en_hoja(ws, ficha):
    """
    Dada una worksheet ws (ya con formato base aplicado),
    escribe en las celdas fijas según la estructura de tu plantilla.
    """

    # 1. PARTE 1 (Encabezado)
    # -------------------------------------------------
    # Celda B5:
    ws["B5"] = ficha["parte1"].get("codigo_ordenado", "")
    # Celda B2: 'objetivo'
    ws["B6"] = ficha["parte1"].get("objetivo", "")

    # Celda B3: 'indicador'
    ws["B7"] = ficha["parte1"].get("indicador", "")

    # 2. METADATOS
    # -------------------------------------------------
    md = ficha.get("metadatos", {})

    # Celda B5: 'definición'
    ws["B9"] = md.get("definicion", "")

    # Celda B6: 'tipo_indicador'
    ws["B11"] = md.get("tipo_indicador", "")

    # Celda B7: 'tipo_algoritmo'
    ws["B10"] = md.get("tipo_algoritmo", "")

    # Celda B8: 'periodicidad'
    ws["B12"] = md.get("periodicidad", "")

    # Celda B9: 'tendencia'
    ws["B13"] = md.get("tendencia", "")

    # Celda B10: 'ambito_medicion'
    ws["B14"] = md.get("ambito_medicion", "")

    # Celda B11: 'dimension_desempeno'
    ws["B15"] = md.get("dimension_desempeno", "")

    # 3. TABLA VARIABLES
    # -------------------------------------------------
    # Las variables pueden ser de 1 a 4. Las filas reservadas son 13, 14, 15 y 16.
    # Su estructura en el diccionario:
    # ficha["variable"] = {
    #     "335": {"name": "...", "medio_verificacion": "..."},
    #     "336": {...}, ...
    # }

    variables = ficha.get("variable", {})
    # Tomamos sólo las keys que sean numéricas para variables reales
    ids_vars = [k for k in variables.keys() if str(k).isdigit()]
    # Ordenamos para mantener consistencia
    ids_vars.sort()

    # Recorremos cada ID y lo escribimos en la fila correspondiente
    fila_var_base = 18
    for idx, var_id in enumerate(ids_vars):
        if idx >= 4:
            break  # máximo 4 variables (335, 336, 337, etc.)

        nombre_var = variables[var_id].get("name", "")
        medio_ver = variables[var_id].get("medio_verificacion", "")

        # Columna b: nombre de la variable
        celda_B = f"B{fila_var_base + idx}"
        ws[celda_B] = nombre_var

        # Columna D: medio de verificación
        celda_D = f"D{fila_var_base + idx}"
        ws[celda_D] = medio_ver

        # (Si en tu plantilla la columna B es para “Nombre” y quieres escribir ahí algo,
        #  podrías usar ws[f"B{fila_var_base + idx}"] = <otro dato>, pero en tu ejemplo
        #  no hay campo extra en 'variable', así que la dejamos en blanco.)

    # 4. TABLA LÍNEA BASE
    # -------------------------------------------------
    # Se arrancan en la fila 24, columnas A (variable), B (valor), C (unidad), D (fecha)
    lb = ficha.get("lineabase", {})

    # Primero, los IDs numéricos (335, 336, 337, …)
    ids_lb = [k for k in lb.keys() if str(k).isdigit()]
    ids_lb.sort()

    fila_lb_base = 24
    for idx, lb_id in enumerate(ids_lb):
        if idx >= 4:
            break  # máximo 4 variables

        # Nombre de la variable: podemos obtenerlo de ficha["variable"][lb_id]["name"]
        nombre_var = ficha.get("variable", {}).get(lb_id, {}).get("name", "")

        valor_lb = lb[lb_id].get("valor", "")
        unidad_lb = lb[lb_id].get("unidad_medida", "")
        fecha_lb = lb[lb_id].get("fecha", "")

        #ws[f"A{fila_lb_base + idx}"] = nombre_var
        ws[f"B{fila_lb_base + idx}"] = valor_lb
        ws[f"C{fila_lb_base + idx}"] = unidad_lb
        ws[f"D{fila_lb_base + idx}"] = fecha_lb

    if "resultado" in lb:
        idx = len(ids_lb)
        if idx < 4:
            if round(valor_lb, 0) == -999:
                valor_lb = 'No Disponible'
                unidad_lb = ''
            elif round(valor_lb, 0) == -998:
                valor_lb = 'No Aplica'
                unidad_lb = ''
            elif round(valor_lb, 0) == -997:
                valor_lb = 'Sin Medición'
                unidad_lb = ''
            else:
                valor_lb  = lb["resultado"].get("valor", "")
                unidad_lb = lb["resultado"].get("unidad_medida", "")
                
            fecha_lb  = lb["resultado"].get("fecha", "")

            #ws[f"A{fila_lb_base + idx}"] = nombre_var
            ws[f"B28"] = valor_lb
            ws[f"C28"] = unidad_lb
            ws[f"D28"] = fecha_lb

    # 5. TABLA META
    # -------------------------------------------------
    # Se arrancan en la fila 31, columnas A (variable), B (valor), C (unidad), D (fecha)
    mt = ficha.get("meta", {})

    # IDs numéricos primero
    ids_mt = [k for k in mt.keys() if str(k).isdigit()]
    ids_mt.sort()

    fila_meta_base = 31
    for idx, mt_id in enumerate(ids_mt):
        if idx >= 4:
            break

        nombre_var = ficha.get("variable", {}).get(mt_id, {}).get("name", "")

        valor_mt = mt[mt_id].get("valor", "")
        unidad_mt = mt[mt_id].get("unidad_medida", "")
        fecha_mt = mt[mt_id].get("fecha", "")

        #ws[f"A{fila_meta_base + idx}"] = nombre_var
        ws[f"B{fila_meta_base + idx}"] = valor_mt
        ws[f"C{fila_meta_base + idx}"] = unidad_mt
        ws[f"D{fila_meta_base + idx}"] = fecha_mt

    # Finalmente, si hay clave "resultado" en meta, la insertamos en la row 35
    if "resultado" in mt:
        idx = len(ids_mt)
        if idx < 4:
            if round(valor_mt, 0) == -999:
                valor_mt = 'No Disponible'
                unidad_mt = ''
            elif round(valor_mt, 0) == -998:
                valor_mt = 'No Aplica'
                unidad_mt = ''
            elif round(valor_mt, 0) == -997:
                valor_mt = 'Sin Medición'
                unidad_mt = ''
            else:
                valor_mt  = mt["resultado"].get("valor", "")
                unidad_mt = mt["resultado"].get("unidad_medida", "")
            
            fecha_mt  = mt["resultado"].get("fecha", "")

            #ws[f"A{fila_meta_base + idx}"] = nombre_var
            ws[f"B35"] = valor_mt
            ws[f"C35"] = unidad_mt
            ws[f"D35"] = fecha_mt

mapeos_predefinidos = {
    'problema': {
        'central': 'problema_central',
        'superior': 'efecto_superior',
        'indirectos': 'efectos_indirectos',
        'directos': 'efectos_directos',
        'cd': 'causas_directas',
        'ci': 'causas_indirectas',
        'id_ed_ei': 'id_EI',
        'id_cd_pc': 'id_PC',
        'id_ci_cd': 'id_CD'
    },
    'solucion': {
        'central': 'solucion_problema',
        'superior': 'fin_superior',
        'indirectos': 'fines_indirectos',
        'directos': 'fines_directos',
        'cd': 'medios_directos',
        'ci': 'medios_indirectos',
        'id_ed_ei': 'id_FI',
        'id_cd_pc': 'id_SP',
        'id_ci_cd': 'id_MD'
    }
}

output_file = 'prueba2.xlsx'
data_map = {
    'Árbol de Problemas': {'problema_central': {'id': 79, 'texto': 'Las niñas y adolescentes entre 10 y 19 años tienen alta incidencia de embarazos'}, 'efecto_superior': {'id': 62, 'texto': 'Reducir las brechas de género en salud'}, 'causas_directas': [{'id': 91, 'texto': 'abusos sexuales contra las nilas y adolescentes', 'id_PC': 79}, {'id': 92, 'texto': 'Uso incorrecto de métodos anticonceptivos en las relaciones sexuales', 'id_PC': 79}, {'id': 93, 'texto': 'Modelo dominantes de ser mujer y madre', 'id_PC': 79}], 'causas_indirectas': [{'id': 153, 'texto': 'Desconocimiento sobre formas de denuncia y atencion', 'id_CD': 91}, {'id': 154, 'texto': 'violencia en la pareja', 'id_CD': 91}, {'id': 155, 'texto': 'violencias en el contexto', 'id_CD': 91}, {'id': 156, 'texto': 'violación y desconocimento de los DDHH', 'id_CD': 91}, {'id': 157, 'texto': 'No hay acceso a servicios e insumos de ssr', 'id_CD': 92}, {'id': 158, 'texto': 'Ausencia de educación integral de la sexualidad', 'id_CD': 92}, {'id': 159, 'texto': 'Deficiente comunicación y psicoefectividad', 'id_CD': 92}, {'id': 160, 'texto': 'Vulneración del estado laico', 'id_CD': 92}, {'id': 161, 'texto': 'Roles tradicionales de género', 'id_CD': 93}, {'id': 162, 'texto': 'baja escolaridad', 'id_CD': 93}, {'id': 163, 'texto': 'falta de acceso al desarrollo y los derechos', 'id_CD': 93}], 'efectos_directos': [{'id': 106, 'texto': 'Trabajos precarios y mal remunerados', 'id_EI': 74, 'id_PC': 79}, {'id': 107, 'texto': 'Bajo rendimiento y deserción escolar', 'id_EI': 74, 'id_PC': 79}, {'id': 108, 'texto': 'Mayor frecuencia de embarazos subsecuentes en madres adolescentes', 'id_EI': 74, 'id_PC': 79}, {'id': 109, 'texto': 'complicaciones en el embarzo, parto y puerperio', 'id_EI': 74, 'id_PC': 79}, {'id': 110, 'texto': 'Interrupción del emabarazo en clandestinidad y condiciones inseguras', 'id_EI': 74, 'id_PC': 79}, {'id': 111, 'texto': 'reproducción del círculo de la pobreza', 'id_EI': 74, 'id_PC': 79}], 'efectos_indirectos': [{'id': 74, 'texto': 'Disminución de desarrollo potencial de capacidades y afectaciones del bienestar integral de las niñas, niños y adolescentes', 'id_ES': 62}]},
    'Árbol de Objetivos': {'solucion_problema': {'id': 8, 'texto': 'Implementar programas de educación sexual integral y acceso a servicios de salud reproductiva para niñas y adolescentes.'}, 'fin_superior': {'id': 8, 'texto': 'Lograr la equidad de género en el acceso y calidad de los servicios de salud.'}, 'medios_directos': [{'id': 17, 'texto': 'Sistema judicial ineficaz.', 'id_SP': 8}, {'id': 18, 'texto': 'Falta de educación sexual adecuada.', 'id_SP': 8}, {'id': 19, 'texto': 'La presión social sobre las mujeres para priorizar la familia sobre sus carreras.', 'id_SP': 8}], 'medios_indirectos': [{'id': 32, 'texto': 'Campañas de difusión masiva sobre procesos de denuncia y atención disponibles.', 'id_MD': 17}, {'id': 33, 'texto': 'Influencias culturales que normalizan comportamientos abusivos.', 'id_MD': 17}, {'id': 34, 'texto': 'Generación de tensiones dentro de las comunidades.', 'id_MD': 17}, {'id': 35, 'texto': 'Desigualdad social y económica.', 'id_MD': 17}, {'id': 36, 'texto': 'Implementar programas de distribución y acceso a servicios e insumos de salud sexual y reproductiva.', 'id_MD': 18}, {'id': 37, 'texto': 'Implementar programas educativos integrales sobre sexualidad en escuelas.', 'id_MD': 18}, {'id': 38, 'texto': 'Mejora del entrenamiento en habilidades de comunicación y gestión emocional.', 'id_MD': 18}, {'id': 39, 'texto': 'Promoción de políticas religiosas en la educación pública.', 'id_MD': 18}, {'id': 40, 'texto': 'La persistencia de los roles tradicionales de género.', 'id_MD': 19}, {'id': 41, 'texto': 'El acceso limitado a educación de calidad afecta el desarrollo de habilidades necesarias para el empleo.', 'id_MD': 19}, {'id': 42, 'texto': 'Promover la educación y el acceso a los derechos básicos.', 'id_MD': 19}], 'fines_directos': [{'id': 15, 'texto': 'Trabajos estables y bien remunerados.', 'id_FI': 11, 'id_SP': 8}, {'id': 16, 'texto': 'Aumento del rendimiento académico y disminución de la deserción escolar.', 'id_FI': 11, 'id_SP': 8}, {'id': 17, 'texto': 'Reducción de la frecuencia de embarazos subsecuentes en madres adolescentes.', 'id_FI': 11, 'id_SP': 8}, {'id': 18, 'texto': 'Reducción de complicaciones en el embarazo, parto y puerperio.', 'id_FI': 11, 'id_SP': 8}, {'id': 19, 'texto': 'Acceso seguro y legal al aborto.', 'id_FI': 11, 'id_SP': 8}, {'id': 20, 'texto': 'Acceso universal a oportunidades equitativas.', 'id_FI': 11, 'id_SP': 8}], 'fines_indirectos': [{'id': 11, 'texto': 'Mejora del desarrollo potencial de capacidades y bienestar integral de las niñas, niños y adolescentes.', 'id_FS': 8}]}
    }
